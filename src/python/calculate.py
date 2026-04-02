import pandas as pd
from openpyxl import load_workbook
import config

def get_best_column_alias(df_filt, df, keyword, prefix_match=False):
    kw = keyword.lower().strip()
    for col in df.columns:
        c = str(col).lower().strip()
        if c == kw or c.startswith(kw + '.') or (prefix_match and c.startswith(kw)):
            return col
    return None

def find_summary_row(ws, label, start_row=1, max_row=1000):
    for row in range(start_row, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None and str(cell_value).strip() == str(label).strip():
            return row
    raise ValueError(
        f"Row label '{label}' not found in column A between rows {start_row} and {max_row}"
    )

def detect_is_empty(ws, target_col):
    check_rows = [4, 6, 29, 30, 32]
    for r in check_rows:
        val = ws.cell(row=r, column=target_col).value
        if val is not None and val != "":
            return False  # data found — ROLLING mode
    return True           # all empty — EXPANSION mode

def shift_columns(wb, is_empty=False):
    ws = wb["Overall Summary"]

    column_to_update = 15   # Column O — the current month slot
    profile_column   = 43   # Column AQ
    unsatisfied_col  = 44   # Column AR

    if is_empty:
        # EXPANSION: source = destination (no shift, just fill col 15 in place)
        main_source_start_col    = 3
        main_source_end_col      = 14
        main_destination_start_col = 3

        prof_src_start  = 19
        prof_src_end    = 42
        prof_dest_start = 19

        like_source_start_col      = 73
        like_source_end_col        = 168
        like_destination_start_col = 73

    else:
        # ROLLING: shift left by 1 column (D→C, E→D … O→N), then clear O
        main_source_start_col    = 4
        main_source_end_col      = 15
        main_destination_start_col = 3

        prof_src_start  = 21
        prof_src_end    = 44
        prof_dest_start = 19

        like_source_start_col      = 81
        like_source_end_col        = 176
        like_destination_start_col = 73

    # We only shift rows 4 onwards — rows 1–3 are headers/labels that never change.
    # The row range goes to 500 to cover all data rows in the sheet.
    row_ranges = [(4, 500)]

    # --- snapshot_block ---
    # Reads a rectangular block of cells into memory.
    # Returns (width, list of (row_index, [cell_values]))
    def snapshot_block(ws, src_start, src_end, row_ranges):
        width = src_end - src_start + 1
        data = []
        for r0, r1 in row_ranges:
            for r in range(r0, r1 + 1):
                row_vals = [
                    ws.cell(row=r, column=c).value
                    for c in range(src_start, src_end + 1)
                ]
                data.append((r, row_vals))
        return width, data

    # --- write_block ---
    # Writes the snapshotted data to a new destination column range.
    def write_block(ws, dst_start, width, snapped_rows):
        for (r, row_vals) in snapped_rows:
            for j in range(width):
                ws.cell(row=r, column=dst_start + j).value = row_vals[j]

    # --- clear_cols ---
    # Clears specific columns within the given row ranges.
    # Used after a ROLLING shift to blank out the vacated rightmost column
    # so the new month's data goes in clean.
    def clear_cols(ws, cols, row_ranges):
        for r0, r1 in row_ranges:
            for r in range(r0, r1 + 1):
                for c in cols:
                    ws.cell(row=r, column=c).value = None

    # Shift all three data blocks
    print("  Shifting main data block...")
    main_width, main_snap = snapshot_block(
        ws, main_source_start_col, main_source_end_col, row_ranges
    )
    write_block(ws, main_destination_start_col, main_width, main_snap)

    print("  Shifting profile history block...")
    prof_width, prof_snap = snapshot_block(
        ws, prof_src_start, prof_src_end, row_ranges
    )
    write_block(ws, prof_dest_start, prof_width, prof_snap)

    print("  Shifting likelihood profiles block...")
    like_width, like_snap = snapshot_block(
        ws, like_source_start_col, like_source_end_col, row_ranges
    )
    write_block(ws, like_destination_start_col, like_width, like_snap)

    # In ROLLING mode, clear the vacated rightmost columns so they're clean
    if not is_empty:
        print("  Clearing vacated columns...")
        cols_to_clear = [column_to_update, profile_column, unsatisfied_col]

        # Also clear the tail end of the likelihood block that was shifted left
        like_tail_start = like_destination_start_col + like_width
        like_tail_end   = like_source_end_col
        if like_tail_start <= like_tail_end:
            cols_to_clear += list(range(like_tail_start, like_tail_end + 1))

        clear_cols(ws, cols_to_clear, row_ranges)


# ---------------------------------------------------------------------------
def calc_satisfaction(df: pd.DataFrame, ws, target_col: int) -> None:
    """
    Calculates overall satisfaction percentage.
    Denominator: All respondents (len(df))
    Adds to 100%: Yes / No
    """
    RAW_COL = "Overall, were you satisfied with your most recent experience at Edinburgh Airport? "
    
    # Bug 1: values are 1/0 integers, not "Yes"/"No"
    count_yes = df[RAW_COL].eq(1).sum()
    count_no = df[RAW_COL].eq(0).sum()
    total = len(df)
    
    pct_yes = count_yes / total if total > 0 else 0
    pct_no = count_no / total if total > 0 else 0
    
    # Use bounds to find the correct "Yes" / "No" under the "Satisfaction " header
    header_row = find_summary_row(ws, "Satisfaction ")
    row_yes = find_summary_row(ws, "Yes", start_row=header_row + 1, max_row=header_row + 5)
    row_no = find_summary_row(ws, "No", start_row=header_row + 1, max_row=header_row + 5)
    
    # Write Yes
    cell_yes = ws.cell(row=row_yes, column=target_col)
    cell_yes.value = float(pct_yes)
    cell_yes.number_format = '0.0%'
    
    # Write No
    cell_no = ws.cell(row=row_no, column=target_col)
    cell_no.value = float(pct_no)
    cell_no.number_format = '0.0%'
    
    print(f"    - satisfaction Yes: {pct_yes:.1%} ({count_yes}/{total})")
    print(f"    - satisfaction No: {pct_no:.1%} ({count_no}/{total})")

def calc_nps(df: pd.DataFrame, ws, target_col: int) -> None:
    """
    Calculates NPS (Net Promoter Score).
    Promoters (9-10) - Detractors (0-6)
    Denominator: Respondents who answered the question (dropna)
    Writes integer value to "NPS score"
    """
    RAW_COL = "How likely are you to recommend Edinburgh Airport to a friend or colleague?"
    
    # Filter to only people who answered
    responses = pd.to_numeric(df[RAW_COL], errors='coerce').dropna()
    total = len(responses)
    
    promoters = (responses >= 9).sum()
    detractors = (responses <= 6).sum()
    
    if total > 0:
        nps = (promoters / total) * 100 - (detractors / total) * 100
        nps = int(round(nps))
    else:
        nps = 0
        
    row = find_summary_row(ws, "NPS score")
    
    cell = ws.cell(row=row, column=target_col)
    cell.value = float(nps)
    cell.number_format = '0'
    
    print(f"    - nps: {nps} (Promoters: {promoters}, Detractors: {detractors}, Total: {total})")

def calc_reason_for_travel(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Why were you travelling? "
    TRAVEL_REASON_MAP = {
        "Business ":  "Business",
        "Leisure ":   "Leisure",
        "VFR ":       "Visiting Family and Friends",
    }
    
    responses = df[RAW_COL].dropna()
    total = len(responses)
    
    for summary_label, raw_val in TRAVEL_REASON_MAP.items():
        count = (responses == raw_val).sum()
        pct = count / total if total > 0 else 0
        
        row = find_summary_row(ws, summary_label)
        cell = ws.cell(row=row, column=target_col)
        cell.value = float(pct)
        cell.number_format = '0.0%'
        print(f"    - {summary_label.strip()}: {pct:.1%} ({count}/{total})")

def calc_journey_type(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Which of the following best describes your recent journey through Edinburgh Airport?"
    
    responses = df[RAW_COL].dropna()
    total = len(responses)
    
    for summary_label in ["I arrived at Edinburgh Airport", "I departed from Edinburgh Airport"]:
        count = (responses == summary_label).sum()
        pct = count / total if total > 0 else 0
        
        row = find_summary_row(ws, summary_label)
        cell = ws.cell(row=row, column=target_col)
        cell.value = float(pct)
        cell.number_format = '0.0%'
        print(f"    - {summary_label}: {pct:.1%} ({count}/{total})")

def calc_age_groups(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "What age group do you fall into?"
    
    responses = df[RAW_COL].dropna()
    
    # Granular raw counts
    granular_labels = ["Under 18", "19 - 25", "26 - 35", "36 - 45", "46 - 55", "56 - 65", "66 - 75", "75+"]
    for label in granular_labels:
        count = (responses == label).sum()
        try:
            row = find_summary_row(ws, label)
            cell = ws.cell(row=row, column=target_col)
            cell.value = int(count)
            cell.number_format = '0'
        except ValueError:
            pass # Ignore if granular label is missing from summary sheet for some reason
            
    # Consolidated percentages
    AGE_REMAP = {
        "Under 18": "<18",
        "19 - 25":  "19-35",  "26 - 35": "19-35",
        "36 - 45":  "36-55",  "46 - 55": "36-55",
        "56 - 65":  "56-65",
        "66 - 75":  "65+",    "75+":     "65+",
    }
    
    age_band = responses.map(AGE_REMAP)
    total = age_band.notna().sum()  # Excludes "Prefer not to say" or unmapped values
    
    consolidated_labels = ["<18", "19-35", "36-55", "56-65", "65+"]
    for label in consolidated_labels:
        count = (age_band == label).sum()
        pct = count / total if total > 0 else 0
        row = find_summary_row(ws, label)
        cell = ws.cell(row=row, column=target_col)
        cell.value = float(pct)
        cell.number_format = '0.0%'
        print(f"    - Age {label}: {pct:.1%} ({count}/{total})")

def calc_gender(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "What is your gender?"
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    
    labels = ["Female", "Male", "Non-binary", "Prefer not to say", "I prefer to self-describe"]
    for label in labels:
        count = (responses == label).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, label)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Gender {label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_transport_mode(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "How did you get to/from the airport?"
    TRANSPORT_MAP = {
        "Drove rental car":              "Drove personal vehicle",
        "Drove rental vehicle":          "Drove rental vehicle",
        "Dropped off by friends/family": "Dropped off by family/friends",
        "Taxi, limo or private hire":    "Taxi/Limousine/Private Hire",
        "Rideshare service":             "Rideshare services (e.g Uber)",
        "Walked/shuttled via airport hotel": "Walked/shuttled via airport hotel",
        "Bus":                           "Bus",
        "Tram":                          "Tram",
        "Bicycle":                       "Bicycle",
        "Flight connection / Layover":   "Flight connection / Layover",
    }
    
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    
    for summary_label, raw_val in TRANSPORT_MAP.items():
        count = (responses == raw_val).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Transport {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass
            
    # Calculate "Other" (anything not in the explicit map)
    known_raw_vals = set(TRANSPORT_MAP.values())
    other_count = (~responses.isin(known_raw_vals)).sum()
    other_pct = other_count / total if total > 0 else 0
    try:
        row = find_summary_row(ws, "Other", start_row=58, max_row=69) # anchor to transport section
        cell = ws.cell(row=row, column=target_col)
        cell.value = float(other_pct)
        cell.number_format = '0.0%'
        print(f"    - Transport Other: {other_pct:.1%} ({other_count}/{total})")
    except ValueError:
        pass

def calc_group_size(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Including yourself how many people were in your travel party? "
    # pandas might convert int to float if there are NaNs, so "1" becomes "1.0", we strip ".0"
    responses = df[RAW_COL].dropna().astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    total = len(responses)
    
    GROUP_MAP = {
        "1": "1",
        "2": "2",
        "3": "3",
        "4": "4",
        "5": "5",
        "6 or more": "6+"
    }
    
    for raw_val, summary_label in GROUP_MAP.items():
        count = (responses == raw_val).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=69, max_row=76)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Group Size {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_travel_frequency(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "How frequently do you travel by air? "
    FREQ_MAP = {
        "Less than once a year":   "Less than once (<1) per year",
        "1 or 2 times a year":     "1 or 2 times a year",
        "3 to 5 times a year":     "3 to 5 times a year",
        "6 to 10 times a year":    "6 to 10 times a year",
        "11 to 20 times a year":   "11 to 20 times a year",
        "20+ times a year":        "more than 20 times a year"
    }
    
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    
    for raw_val, summary_label in FREQ_MAP.items():
        count = (responses == raw_val).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=76, max_row=85)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Travel Freq {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_arrival_time(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "How long before your flight did you arrive at the airport?"
    ARRIVAL_MAP = {
        "&lt;60 minutes":                        "Less than 60 minutes before take-off",
        "<60 minutes":                           "Less than 60 minutes before take-off",
        "Less than 60 minutes before take-off":  "Less than 60 minutes before take-off",
        
        "60 - 120 minutes":                      "60 - 120 minutes before take-off",
        "60 - 120 minutes before take-off":      "60 - 120 minutes before take-off",
        
        "120 - 180 minutes":                     "120 - 180 minutes before take-off",
        "120 - 180 minutes before take-off":     "120 - 180 minutes before take-off",
        
        "120+ minutes":                          "More than 180 minutes before take-off",
        "More than 180 minutes before take-off": "More than 180 minutes before take-off"
    }
    
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    
    # Consolidate counts: multiple raw values can map to same summary label
    summary_counts = {}
    for raw_val, summary_label in ARRIVAL_MAP.items():
        
        count = (responses == raw_val).sum()
        summary_counts[summary_label] = summary_counts.get(summary_label, 0) + count
        
    for summary_label, count in summary_counts.items():
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=89, max_row=95)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Arrival {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_religion(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "What is your religion or belief?"
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    
    labels = ["Buddhist", "Christian", "Hindu", "Jewish", "Muslim", "No religion", "Other", "Prefer not to say", "Sikh"]
    for label in labels:
        count = (responses == label).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, label, start_row=338, max_row=350)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Religion {label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_coaching_satisfaction(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "I was satisfied with my coaching experience"
    if RAW_COL not in df.columns:
        return
        
    responses = df[RAW_COL].dropna().str.lower().str.strip()
    total = len(responses)
    
    labels = ["Strongly agree", "Agree", "Somewhat agree", "Somewhat disagree", "Disagree", "Strongly disagree"]
    
    for label in labels:
        count = (responses == label.lower()).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, label, start_row=190, max_row=200)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Coaching {label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_travel_companions(df: pd.DataFrame, ws, target_col: int) -> None:
    labels = [
        "With partner", 
        "With friend(s)", 
        "With children ages 0-2", 
        "With children ages 3-9", 
        "With children ages 10-17", 
        "With other relative(s)", 
        "With colleague(s)"
    ]
    
    total = len(df) # Multi-select, denominator is all respondents
    for label in labels:
        if label in df.columns:
            count = df[label].notna().sum()
            pct = count / total if total > 0 else 0
            try:
                row = find_summary_row(ws, label, start_row=49, max_row=60)
                cell = ws.cell(row=row, column=target_col)
                cell.value = float(pct)
                cell.number_format = '0.0%'
                print(f"    - Companion {label}: {pct:.1%} ({count}/{total})")
            except ValueError:
                pass

def calc_scotland_stay_region(df: pd.DataFrame, ws, target_col: int) -> None:
    labels = [
        "Edinburgh",
        "Glasgow",
        "Fife & Kindross",
        "Borders",
        "Highlands & Islands",
        "I did not stay in Scotland"
    ]
    
    existing_cols = [col for col in labels if col in df.columns]
    if not existing_cols:
        return
        
    answered_mask = df[existing_cols].notna().any(axis=1)
    total = answered_mask.sum()
    
    for label in existing_cols:
        count = df[label].notna().sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, label, start_row=112, max_row=120)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Scotland Stay {label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_why_travel_access_hard(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "It was easy to get to / from the airport"
    if BASE_COL not in df.columns:
        return
        
    # Filter first: Respondents who disagreed with access question
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty:
        return
        
    LABEL_MAP = {
        "Public transport links were tricky, limited or confusing": "Public transport links were tricky, limited or confusing",
        "Roadworks, traffic or congestion": "Roadworks, traffic or congestion",
        "Signage and direction to the airport was not clear": "Signage and direction to the airport was not clear",
        "Drop-off was expensive": "Drop-off was expensive"
    }
    
    best_something = get_best_column_alias(df_filt, df, "something else")
    if best_something:
        LABEL_MAP[best_something] = "Something else"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols:
        return
        
    # Denominator = total selections across all why-options (not respondent count)
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=171, max_row=175)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Access Hard {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_wayfinding_hard(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "It was easy to find my way around the airport"
    if BASE_COL not in df.columns: return
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Not enough signs": "Not enough signs",
        "Signs were confusing": "Signs were confusing",
        "Too many signs": "Too many signs",
        "No one around to direct me": "No one around to direct me"
    }
    best_something = get_best_column_alias(df_filt, df, "something else")
    if best_something:
        LABEL_MAP[best_something] = "Something else"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=212, max_row=216)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Wayfinding {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_which_staff_rude(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "I was satisfied with the staff service I received"
    if BASE_COL not in df.columns: return
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Check-in hall staff": "Check-in hall staff",
        "Special assistance staff": "Special assistance staff",
        "Security staff": "Security staff",
        "Gate staff": "Gate staff",
        "Airline staff": "Airline staff",
        "Immigration hall staff": "Immigration hall staff",
        "Baggage reclaim staff": "Baggage reclaim staff",
        "Restaurant/shop staff": "Restaurant/shop staff",
        "Lounge staff": "Lounge staff"
    }
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=230, max_row=238)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Rude {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_coaching_unsatisfying(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "I was satisfied with my coaching experience"
    if BASE_COL not in df.columns: return
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "The coach ride was too long": "The coach ride was too long",
        "I waited a long time to get off the coach": "I waited a long time to get off the coach",
        "There were too many people on the coach": "There were too many people on the coach",
        "I was not satisfied with the staff service": "I was not satisfied with the staff service"
    }
    best_other = get_best_column_alias(df_filt, df, "Other")
    if best_other:
        LABEL_MAP[best_other] = "Other"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=199, max_row=203)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Coach Unsat {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_no_food_purchased(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Why didn’t you buy any food or drink?"
    if RAW_COL not in df.columns: return
    
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    if total == 0: return

    for label, count in responses.value_counts().items():
        pct = count / total
        try:
            row = find_summary_row(ws, label, start_row=148, max_row=153)
            ws.cell(row=row, column=target_col).value = float(pct)
            ws.cell(row=row, column=target_col).number_format = '0.0%'
            print(f"    - Food No {label[:15]}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_wifi_poor_coverage_locations(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = get_best_column_alias(df, df, "poor coverage in certain areas")
    if not BASE_COL: return
    df_filt = df[df[BASE_COL].notna()]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Baggage hall": "Baggage hall",
        "Check-in hall": "Check-in hall",
        "Security hall": "Security hall",
        "Departure lounge": "Departure lounge",
        "Restaurant / Bar / Coffee shop": "Restaurant / Bar / Coffee shop",
        "Gate area": "Gate area",
        "Immigration hall": "Immigration hall",
        "Car parks": "Car parks",
        "Retail store": "Retail store"
    }
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=297, max_row=305)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Wifi Coverage {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass
            
def calc_food_purchase_location(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "Did you purchase any food or drink at Edinburgh Airport?"
    if BASE_COL not in df.columns: return
    # 1 = Yes
    df_filt = df[df[BASE_COL] == 1]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Before security (departures)": "Before security (departures)",
        "After security (departures)": "After security (departures)",
        "After landing (arrivals)": "After landing (arrivals)"
    }
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=144, max_row=146)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Food Location {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_agreement_questions(df: pd.DataFrame, ws, target_col: int) -> None:
    AGREEMENT_MAP = {
        "It was easy to get to/from the airport": (164, 169),
        "It was easy to find information about my flight": (177, 182),
        "It was easy to find my way around the airport": (205, 210),
        "I was satisfied with staff service at the airport ": (218, 223),
        "I was satisfied with the cleanliness of the airport": (241, 246),
        "It was easy to find a seat in the airport": (255, 260),
        "It was easy to recycle at the airport ": (269, 274),
        "I was satisfied with the Wi-Fi service at the airport": (282, 287)
    }
    
    likert_labels = ["Strongly agree", "Agree", "Somewhat agree", "Somewhat disagree", "Disagree", "Strongly disagree"]
    
    for raw_col, (start_r, end_r) in AGREEMENT_MAP.items():
        if raw_col not in df.columns:
            continue
            
        responses = df[raw_col].dropna().str.lower().str.strip()
        total = len(responses)
        if total == 0:
            continue
            
        for label in likert_labels:
            count = (responses == label.lower()).sum()
            pct = count / total
            try:
                row = find_summary_row(ws, label, start_row=start_r, max_row=end_r)
                cell = ws.cell(row=row, column=target_col)
                cell.value = float(pct)
                cell.number_format = '0.0%'
            except ValueError:
                pass
                
        print(f"    - Agreement [{raw_col}] calculated ({total} respondents)")

def calc_geographic_split(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "What country do you live in? "
    if RAW_COL not in df.columns: return

    responses = df[RAW_COL].dropna().astype(str).str.strip().str.title()
    total = len(responses)

    scotland_count = (responses == "Scotland").sum()
    
    # Rest of UK logic
    uk_terms = ["England", "Wales", "Northern Ireland", "United Kingdom", "Uk", "Great Britain", "Gb"]
    rest_of_uk_count = responses.isin(uk_terms).sum()
    
    # Rest of World logic
    rest_of_world_count = total - (scotland_count + rest_of_uk_count)

    split_map = {
        "Scotland": scotland_count,
        "Rest of UK": rest_of_uk_count,
        "Rest of World": rest_of_world_count
    }

    for summary_label, count in split_map.items():
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=102, max_row=104)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Geo Split {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_checked_luggage(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Did you check in a bag?"
    if RAW_COL not in df.columns: return

    responses = df[RAW_COL].dropna().astype(str).str.strip().str.title()
    total = len(responses)

    for summary_label in ["Yes", "No"]:
        count = (responses == summary_label).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=36, max_row=39)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Luggage {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_assisted_travel(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Did you use Assisted Travel for your most recent journey?"
    if RAW_COL not in df.columns: return

    responses = df[RAW_COL].dropna().astype(str).str.strip().str.title()
    total = len(responses)

    for summary_label in ["Yes", "No"]:
        count = (responses == summary_label).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=355, max_row=358)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Assisted {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_utm_source_counts(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "utm_source"
    if RAW_COL not in df.columns: return

    responses = df[RAW_COL].dropna().astype(str).str.strip()
    
    UTM_LABELS = [
        "car_parking", "EAL_homepage", "ed-wifi", 
        "welcome_back_newsletter", "wifi_landing_D", 
        "trial_campaign", "wifi_signup"
    ]

    for label in UTM_LABELS:
        count = (responses == label).sum()
        try:
            row = find_summary_row(ws, label, start_row=10, max_row=18)
            ws.cell(row=row, column=target_col).value = int(count)
            print(f"    - UTM [{label}]: {count}")
        except ValueError:
            pass

def calc_environmental_actions(df: pd.DataFrame, ws, target_col: int) -> None:
    ENV_MAP = {
        "Yes, I used public transport, ride shared, cycled or walked": "Yes, I used public transport, ride shared, cycled or walked",
        "Yes, I offset my carbon emissions": "Yes, I offset my carbon emissions",
        "Yes, I brought a reusable cup": "Yes, I brought a reusable cup",
        "No, I didn't take specific actions": "No, I didn't take specific actions",
        "Other.1": "Other"
    }

    total_surveys = len(df)

    for raw_col, display_label in ENV_MAP.items():
        if raw_col in df.columns:
            count = df[raw_col].notna().sum()
            pct = count / total_surveys if total_surveys > 0 else 0
            try:
                row = find_summary_row(ws, display_label, start_row=120, max_row=127)
                cell = ws.cell(row=row, column=target_col)
                cell.value = float(pct)
                cell.number_format = '0.0%'
                print(f"    - Env Action [{display_label}]: {pct:.1%} ({count}/{total_surveys})")
            except ValueError:
                pass

def calc_why_flight_info_hard(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "It was easy to find information about my flight"
    if BASE_COL not in df.columns: return
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Not enough announcements": "Not enough announcements",
        "Announcements were hard to hear": "Announcements were hard to hear",
        "Could not find flight information": "Could not find flight information",
        "Flight information changed too often or came too late": "Flight information changed too often or came too late",
        "Flight screen did not have correct information": "Flight screen did not have correct information",
        "No one around to help": "No one around to help"
    }
    best_something = get_best_column_alias(df_filt, df, "something else")
    if best_something:
        LABEL_MAP[best_something] = "Something else"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=183, max_row=191)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Flight Info {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_cleanliness_poor(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "I was satisfied with the cleanliness of the airport"
    if BASE_COL not in df.columns: return
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Litter": "Litter",
        "Bins not emptied": "Bins not emptied",
        "Dirty washrooms": "Dirty washrooms",
        "General grubbiness": "General grubbiness",
        "Foul odour": "Foul odour"
    }
    
    best_something = get_best_column_alias(df_filt, df, "something else")
    if best_something:
        LABEL_MAP[best_something] = "Something else"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=247, max_row=254)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Cleanliness poor {summary_label}: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_seating_poor(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COL = "It was easy to find a seat in the airport"
    if BASE_COL not in df.columns: return
    df_filt = df[df[BASE_COL].str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Not enough seats before security": "Not enough seats before Security",
        "Not enough seats in baggage hall": "Not enough seats in baggage hall",
        "Seats were uncomfortable": "Seats were uncomfortable",
        "Not enough seats at restaurants": "Not enough seats at restaurants",
        "Not enough seats at gates": "Not enough seats at gates"
    }
    
    best_something = get_best_column_alias(df_filt, df, "other", prefix_match=True)
    if not best_something:
        best_something = get_best_column_alias(df_filt, df, "something else")
    if best_something:
        LABEL_MAP[best_something] = "Something else"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            # We search between rows 261 to 268 for stability
            row = find_summary_row(ws, summary_label, start_row=261, max_row=268)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Seating poor {summary_label[:15]}...: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_recycling_hard(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COLS = [
        "It was easy to recycle at the airport",
        "It was easy to recycle at the airport ",
        "It was easy to recycle at Edinburgh Airport"
    ]
    base_col = next((c for c in BASE_COLS if c in df.columns), None)
    if not base_col: return
    
    df_filt = df[df[base_col].astype(str).str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
        
    LABEL_MAP = {
        "Couldn't find recycling bin": "Couldn't find recycling bin",
        "Bins were too full or overflowing": "Bins were too full or overflowing",
        "Bin labelling unclear": "Bin labelling was unclear",
        "Not enough waste separation": "Not enough waste separation"
    }
    
    best_something = get_best_column_alias(df_filt, df, "other", prefix_match=True)
    if not best_something:
        best_something = get_best_column_alias(df_filt, df, "something else")
    if best_something:
        LABEL_MAP[best_something] = "Something else"
        
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            # We search between rows 275 to 280
            row = find_summary_row(ws, summary_label, start_row=275, max_row=281)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Recycling hard {summary_label[:15]}...: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_wifi_poor(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COLS = [
        "I was satisfied with the WiFi service at Edinburgh Airport",
        "I was satisfied with the WiFi at Edinburgh Airport",
        "I was satisfied with the Wi-Fi service at the airport"
    ]
    base_col = next((c for c in BASE_COLS if c in df.columns), None)
    if not base_col: return
    
    df_filt = df[df[base_col].astype(str).str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
    
    LABEL_MAP = {
        "Unstable connection": "Unstable connection",
        "Didn't want to provide my email address to connect": "Didn't want to provide my email address to connect",
        "Couldn't connect to the Wi-Fi": "Couldn't connect to Wi-Fi",
        "poor coverage in certain areas of the terminal*": "poor coverage in certain areas of the terminal*",
        "Slow connection": "Slow connection",
        "Wi-Fi didn't work with some apps/services (e.g video calling)": "Wi-Fi didn't work with some apps or services"
    }
    
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    total_selections = df_filt[actual_cols].notna().sum().sum()
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df_filt[col].notna().sum()
        pct = count / total_selections if total_selections > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=288, max_row=296)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - WiFi poor {summary_label[:15]}...: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_why_staff_unsatisfying(df: pd.DataFrame, ws, target_col: int) -> None:
    BASE_COLS = [
        "I was satisfied with the staff service at the airport",
        "I was satisfied with staff service at the airport ",
        "I was satisfied with staff service at the airport"
    ]
    base_col = next((c for c in BASE_COLS if c in df.columns), None)
    if not base_col: return
    
    df_filt = df[df[base_col].astype(str).str.lower().str.strip().isin(["disagree", "strongly disagree"])]
    if df_filt.empty: return
def calc_which_staff_rude(df: pd.DataFrame, ws, target_col: int) -> None:
    LABEL_MAP = {
        "Check-in hall staff": "Check-in hall staff",
        "Special assistance staff": "Special assistance staff",
        "Security staff": "Security staff",
        "Gate staff": "Gate staff",
        "Airline staff": "Airline staff",
        "Immigration hall staff": "Immigration hall staff",
        "Baggage reclaim staff": "Baggage reclaim staff",
        "Restaurant/shop staff": "Restaurant/shop staff",
        "Lounge staff": "Lounge staff"
    }
    
    actual_cols = [c for c in LABEL_MAP.keys() if c in df.columns]
    if not actual_cols: return
    
    total_selections = df[actual_cols].notna().sum().sum()
    if total_selections == 0: return
    
    for col in actual_cols:
        summary_label = LABEL_MAP[col]
        count = df[col].notna().sum()
        pct = count / total_selections
        try:
            row = find_summary_row(ws, summary_label, start_row=229, max_row=240)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Staff rude {summary_label[:15]}...: {pct:.1%} ({count}/{total_selections})")
        except ValueError:
            pass

def calc_disembarkation_method(df: pd.DataFrame, ws, target_col: int) -> None:
    # Future-proof logic for the new metric coming next month
    RAW_COL = "How did you board or disembark from your aircraft?"
    if RAW_COL not in df.columns: return
    
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    
    # Map exact strings expected from the template
    OPTIONS = [
        "I used an airbridge / jet bridge (walked directly between the aircraft and terminal)",
        "I walked across the apron between the aircraft and terminal (no coach)",
        "I took an airside bus between the aircraft and terminal",
        "Don’t know / can’t remember"
    ]
    
    for label in OPTIONS:
        # Some encoding issues with apostrophes might happen in raw data, let's fix just in case
        clean_label = label.replace('’', "'")
        
        # Match against raw data (handling both apostrophe styles)
        count = ((responses == label) | (responses == clean_label)).sum()
        pct = count / total if total > 0 else 0
        
        try:
            row = find_summary_row(ws, label, start_row=134, max_row=140)
            cell = ws.cell(row=row, column=target_col)
            cell.value = float(pct)
            cell.number_format = '0.0%'
            print(f"    - Disembarkation {label[:15]}...: {pct:.1%} ({count}/{total})")
        except ValueError:
            pass

def calc_coaching_by_journey_type(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Were you coached to/from your flight?"
    if RAW_COL not in df.columns: return

    responses = pd.to_numeric(df[RAW_COL], errors='coerce').dropna()
    total = len(responses)

    count_yes = (responses == 1).sum()
    count_no = (responses == 0).sum()

    if total > 0:
        pct_yes = count_yes / total
        pct_no = count_no / total
    else:
        pct_yes = pct_no = 0

    try:
        row_yes = find_summary_row(ws, "Yes", start_row=134, max_row=137)
        cell = ws.cell(row=row_yes, column=target_col)
        cell.value = float(pct_yes)
        cell.number_format = '0.0%'
        print(f"    - Coached Yes: {pct_yes:.1%} ({count_yes}/{total})")
    except ValueError: pass

    try:
        row_no = find_summary_row(ws, "No", start_row=134, max_row=137)
        cell = ws.cell(row=row_no, column=target_col)
        cell.value = float(pct_no)
        cell.number_format = '0.0%'
        print(f"    - Coached No: {pct_no:.1%} ({count_no}/{total})")
    except ValueError: pass

def top5_postcodes(df: pd.DataFrame, ws, target_col: int) -> None:
    col = "What is the start of your postcode?"
    if col not in df.columns: return
    
    s = df[col].dropna().astype(str).str.upper().str.strip()
    s = s[(s != "NAN") & (s != "None") & (s != "")]
    if s.empty: return
    
    top5 = s.value_counts(normalize=True).head(5)
    for i, (pc, pct) in enumerate(top5.items(), start=84):
        val = f"{pc} - {pct:.0%}"
        ws.cell(row=i, column=target_col).value = val
        print(f"    - Postcode {i-83}: {val}")

def calc_top5_countries(df: pd.DataFrame, ws, target_col: int) -> None:
    col = "What country do you live in? "
    if col not in df.columns: return
    
    s = df[col].dropna().astype(str).str.strip()
    if s.empty: return
    
    counts = s.value_counts(normalize=True)
    top5 = counts.head(5)
    other_pct = counts.iloc[5:].sum() if len(counts) > 5 else 0
    
    for i, (country, pct) in enumerate(top5.items(), start=95):
        val = f"{country} - {pct:.0%}"
        ws.cell(row=i, column=target_col).value = val
        print(f"    - Country {i-94}: {val}")
        
    other_val = f"Other - {other_pct:.0%}"
    ws.cell(row=100, column=target_col).value = other_val
    print(f"    - Country Other: {other_val}")

# ---------------------------------------------------------------------------
# QUESTION_REGISTRY
# ---------------------------------------------------------------------------
# Each entry is one metric block. The main loop below calls every function
# in order. To add a new question: write a calc_ function and add one line here.
# To remove a question: comment out its entry. The function can stay.
# ---------------------------------------------------------------------------
def calc_top5_airlines(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL_OPTIONS = [
        "Which airline did you travel with?",
        "Which airline are you travelling with?"
    ]
    raw_col = next((c for c in RAW_COL_OPTIONS if c in df.columns), None)
    if not raw_col: return

    responses = df[raw_col].dropna().astype(str).str.strip()
    total = len(responses)
    if total == 0: return

    counts = responses.value_counts()
    top5 = counts.head(5)
    other_count = counts.iloc[5:].sum() if len(counts) > 5 else 0

    START_ROW = 106
    for i, (airline, count) in enumerate(top5.items()):
        row = START_ROW + i
        pct = count / total
        val = f"{pct:.0%} - {airline}"
        ws.cell(row=row, column=target_col).value = val
        print(f"    - Airline {i+1}: {val}")

    # Row 111 = Other
    other_pct = other_count / total if total > 0 else 0
    val_other = f"{other_pct:.0%} - Other"
    ws.cell(row=111, column=target_col).value = val_other
    print(f"    - Airline Other: {val_other}")

def calc_top5_destinations(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL_OPTIONS = [
        "What country did you travel to/from?",
        "What country are you travelling to/from?"
    ]
    raw_col = next((c for c in RAW_COL_OPTIONS if c in df.columns), None)
    if not raw_col: return

    responses = df[raw_col].dropna().astype(str).str.strip()
    total = len(responses)
    if total == 0: return

    counts = responses.value_counts()
    top5 = counts.head(5)
    other_count = counts.iloc[5:].sum() if len(counts) > 5 else 0

    START_ROW = 308
    for i, (dest, count) in enumerate(top5.items()):
        row = START_ROW + i
        pct = count / total
        val = f"{pct:.0%} - {dest}"
        ws.cell(row=row, column=target_col).value = val
        print(f"    - Destination {i+1}: {val}")

    # Row 313 = Other
    other_pct = other_count / total if total > 0 else 0
    val_other = f"{other_pct:.0%} - Other"
    ws.cell(row=313, column=target_col).value = val_other
    print(f"    - Destination Other: {val_other}")

def calc_carbon_credit_type(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "What type of carbon credit did you use to offset your journey"
    if RAW_COL not in df.columns: return
    OFFSET_COL = "Yes, I offset my carbon emissions"
    df_filt = df[df[OFFSET_COL].notna()] if OFFSET_COL in df.columns else df[df[RAW_COL].notna()]
    if df_filt.empty: return
    responses = df_filt[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    if total == 0: return
    for label, count in responses.value_counts().items():
        pct = count / total
        try:
            row = find_summary_row(ws, label, start_row=128, max_row=133)
            ws.cell(row=row, column=target_col).value = float(pct)
            ws.cell(row=row, column=target_col).number_format = '0.0%'
            print(f"    - Carbon Credit {label[:15]}: {pct:.1%} ({count}/{total})")
        except ValueError: pass

def calc_airport_hotel(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Did you stay at an airport hotel?"
    if RAW_COL not in df.columns: return
    responses = df[RAW_COL].dropna().astype(str).str.strip().str.title()
    total = len(responses)
    for summary_label in ["Yes", "No"]:
        count = (responses == summary_label).sum()
        pct = count / total if total > 0 else 0
        try:
            row = find_summary_row(ws, summary_label, start_row=138, max_row=139)
            ws.cell(row=row, column=target_col).value = float(pct)
            ws.cell(row=row, column=target_col).number_format = '0.0%'
            print(f"    - Airport Hotel {summary_label}: {pct:.1%} ({count}/{total})")
        except ValueError: pass

def calc_food_purchase(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Did you purchase any food or drink at Edinburgh Airport?"
    if RAW_COL not in df.columns: return
    count_yes = (df[RAW_COL] == 1).sum()
    count_no = (df[RAW_COL] == 0).sum()
    total = count_yes + count_no
    pct_yes = count_yes / total if total > 0 else 0
    pct_no = count_no / total if total > 0 else 0
    try:
        row_yes = find_summary_row(ws, "Yes", start_row=141, max_row=142)
        ws.cell(row=row_yes, column=target_col).value = float(pct_yes)
        ws.cell(row=row_yes, column=target_col).number_format = '0.0%'
        print(f"    - Food Purchase Yes: {pct_yes:.1%} ({count_yes}/{total})")
    except ValueError: pass
    try:
        row_no = find_summary_row(ws, "No", start_row=141, max_row=142)
        ws.cell(row=row_no, column=target_col).value = float(pct_no)
        ws.cell(row=row_no, column=target_col).number_format = '0.0%'
        print(f"    - Food Purchase No: {pct_no:.1%} ({count_no}/{total})")
    except ValueError: pass

def calc_products_purchased(df: pd.DataFrame, ws, target_col: int) -> None:
    LABELS = [
        "Beauty or skincare products", "Books or magazines", "Clothing or accessories",
        "Duty Free Alcohol / Liquor", "Electronics", "Gifts or souvenirs",
        "Travel essentials (e.g. toiletries, chargers)"
    ]
    total = len(df)
    for label in LABELS:
        if label in df.columns:
            count = df[label].notna().sum()
            pct = count / total if total > 0 else 0
            try:
                row = find_summary_row(ws, label, start_row=154, max_row=162)
                ws.cell(row=row, column=target_col).value = float(pct)
                ws.cell(row=row, column=target_col).number_format = '0.0%'
                print(f"    - Products {label[:15]}: {pct:.1%} ({count}/{total})")
            except ValueError: pass

def calc_fb_options_wanted(df: pd.DataFrame, ws, target_col: int) -> None:
    MAP = {
        "Family Restaurant": "Family resaurant",
        "Grab & Go": "Grab & Go",
        "Premium Bar": "Premium bar",
        "Quick Service Restaurant": "Quick service restaurant",
        "Food Market": "Food Market",
        "Scottish Cuisine": "Scottish Cuisine"
    }
    total = len(df)
    for raw_col, summary_label in MAP.items():
        if raw_col in df.columns:
            count = df[raw_col].notna().sum()
            pct = count / total if total > 0 else 0
            try:
                row = find_summary_row(ws, summary_label, start_row=360, max_row=366)
                ws.cell(row=row, column=target_col).value = float(pct)
                ws.cell(row=row, column=target_col).number_format = '0.0%'
                print(f"    - F&B Wanted {summary_label}: {pct:.1%} ({count}/{total})")
            except ValueError: pass

def calc_transgender(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Do you consider yourself transgender?"
    if RAW_COL not in df.columns: return
    responses = df[RAW_COL].dropna().astype(str).str.strip().str.title()
    total = len(responses)
    for summary_label in ["Yes", "No", "Prefer Not To Say"]:
        count = (responses == summary_label).sum()
        pct = count / total if total > 0 else 0
        try:
            search_label = "Prefer not to say" if summary_label == "Prefer Not To Say" else summary_label
            row = find_summary_row(ws, search_label, start_row=322, max_row=324)
            ws.cell(row=row, column=target_col).value = float(pct)
            ws.cell(row=row, column=target_col).number_format = '0.0%'
            print(f"    - Transgender {search_label}: {pct:.1%} ({count}/{total})")
        except ValueError: pass

def calc_disability(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "Do you consider yourself to have a disability or health condition?"
    if RAW_COL not in df.columns: return
    responses = df[RAW_COL].dropna().astype(str).str.strip().str.title()
    total = len(responses)
    for summary_label in ["Yes", "No", "Prefer Not To Say"]:
        count = (responses == summary_label).sum()
        pct = count / total if total > 0 else 0
        try:
            search_label = "Prefer not to say" if summary_label == "Prefer Not To Say" else summary_label
            row = find_summary_row(ws, search_label, start_row=326, max_row=328)
            ws.cell(row=row, column=target_col).value = float(pct)
            ws.cell(row=row, column=target_col).number_format = '0.0%'
            print(f"    - Disability {search_label}: {pct:.1%} ({count}/{total})")
        except ValueError: pass

def calc_holiday_spend(df: pd.DataFrame, ws, target_col: int) -> None:
    RAW_COL = "On average, how much does your household spend on holidays every year?"
    if RAW_COL not in df.columns: return
    responses = df[RAW_COL].dropna().astype(str).str.strip()
    total = len(responses)
    if total == 0: return
    for label, count in responses.value_counts().items():
        pct = count / total
        try:
            row = find_summary_row(ws, label, start_row=349, max_row=354)
            ws.cell(row=row, column=target_col).value = float(pct)
            ws.cell(row=row, column=target_col).number_format = '0.0%'
            print(f"    - Holiday Spend {label[:10]}...: {pct:.1%} ({count}/{total})")
        except ValueError: pass

QUESTION_REGISTRY = [
    {"name": "satisfaction",      "fn": calc_satisfaction},
    {"name": "nps",               "fn": calc_nps},
    {"name": "reason_for_travel", "fn": calc_reason_for_travel},
    {"name": "journey_type",      "fn": calc_journey_type},
    {"name": "age_groups",        "fn": calc_age_groups},
    {"name": "gender",            "fn": calc_gender},
    {"name": "transport_mode",    "fn": calc_transport_mode},
    {"name": "group_size",        "fn": calc_group_size},
    {"name": "travel_frequency",  "fn": calc_travel_frequency},
    {"name": "arrival_time",      "fn": calc_arrival_time},
    {"name": "religion",          "fn": calc_religion},
    {"name": "coaching_satisfaction", "fn": calc_coaching_satisfaction},
    {"name": "travel_companions", "fn": calc_travel_companions},
    {"name": "scotland_stay_region", "fn": calc_scotland_stay_region},
    {"name": "why_travel_access_hard", "fn": calc_why_travel_access_hard},
    {"name": "why_wayfinding_hard", "fn": calc_why_wayfinding_hard},
    {"name": "which_staff_rude", "fn": calc_which_staff_rude},
    {"name": "why_coaching_unsatisfying", "fn": calc_why_coaching_unsatisfying},
    {"name": "why_no_food_purchased", "fn": calc_why_no_food_purchased},
    {"name": "wifi_poor_coverage_locations", "fn": calc_wifi_poor_coverage_locations},
    {"name": "food_purchase_location", "fn": calc_food_purchase_location},
    {"name": "agreement_questions", "fn": calc_agreement_questions},
    {"name": "geographic_split", "fn": calc_geographic_split},
    {"name": "checked_luggage", "fn": calc_checked_luggage},
    {"name": "assisted_travel", "fn": calc_assisted_travel},
    {"name": "utm_source_counts", "fn": calc_utm_source_counts},
    {"name": "environmental_actions", "fn": calc_environmental_actions},
    {"name": "coaching_by_journey_type", "fn": calc_coaching_by_journey_type},
    {"name": "disembarkation_method", "fn": calc_disembarkation_method},
    {"name": "why_flight_info_hard", "fn": calc_why_flight_info_hard},
    {"name": "why_cleanliness_poor", "fn": calc_why_cleanliness_poor},
    {"name": "why_seating_poor", "fn": calc_why_seating_poor},
    {"name": "why_recycling_hard", "fn": calc_why_recycling_hard},
    {"name": "why_wifi_poor", "fn": calc_why_wifi_poor},
    {"name": "why_staff_unsatisfying", "fn": calc_why_staff_unsatisfying},
    {"name": "top5_postcodes", "fn": top5_postcodes},
    {"name": "top5_countries", "fn": calc_top5_countries},
    {"name": "top5_airlines", "fn": calc_top5_airlines},
    {"name": "top5_destinations", "fn": calc_top5_destinations},
    {"name": "carbon_credit_type", "fn": calc_carbon_credit_type},
    {"name": "airport_hotel", "fn": calc_airport_hotel},
    {"name": "food_purchase", "fn": calc_food_purchase},
    {"name": "products_purchased", "fn": calc_products_purchased},
    {"name": "fb_options_wanted", "fn": calc_fb_options_wanted},
    {"name": "transgender", "fn": calc_transgender},
    {"name": "disability", "fn": calc_disability},
    {"name": "holiday_spend", "fn": calc_holiday_spend},
]


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
if __name__ == "__main__":

    # --- Load workbook ---
    # IMPORTANT: do NOT use data_only=True here.
    # data_only=True reads formula cells as their last cached value.
    # If we then save the file, every formula gets permanently replaced with
    # a plain number — destroying =TODAY(), =TEXT(EOMONTH(...)), =SUM() etc.
    # With data_only=False (the default), formula cells have value=None.
    # The shift logic copies None→None for those cells, which is harmless.
    print(f"Loading workbook from {config.INPUT_PATH}...")
    wb = load_workbook(config.INPUT_PATH)
    ws_summary = wb["Overall Summary"]

    # --- Validate target column ---
    # Row 2 col 15 is a hardcoded literal "Feb-26" (not a formula).
    # Most other row 2 cells are =TEXT(EOMONTH(...)) formulas — they return
    # None under data_only=False, so we can only reliably check col 15.
    # If this check fails, stop immediately — writing to the wrong column
    # would silently corrupt the entire historical dataset.
    target_col = config.TARGET_COLUMN
    label = ws_summary.cell(row=2, column=target_col).value
    if label != config.REPORT_MONTH_LABEL:
        raise ValueError(
            f"Column {target_col} has label '{label}', "
            f"expected '{config.REPORT_MONTH_LABEL}'. "
            f"Check TARGET_COLUMN in config.py."
        )
    print(f"Target column validated: col {target_col} = '{label}'")

    # --- Detect shift mode ---
    # If the target column already has data, we're in ROLLING mode (shift left).
    # If it's empty, we're in EXPANSION mode (just write in place).
    is_empty = detect_is_empty(ws_summary, target_col)
    mode = "EXPANSION" if is_empty else "ROLLING"
    print(f"Shift mode: {mode}")

    # --- Shift columns ---
    print("Shifting columns...")
    shift_columns(wb, is_empty=is_empty)

    # --- Load raw data ---
    # sheet_name=0 always reads the first sheet regardless of tab name.
    # The Typeform export tab is named with a random ID (e.g. "eEHJ9Yfm")
    # that could change if the form is ever recreated. Using index 0 is safer.
    print(f"Loading raw data from {config.RAW_DATA_PATH}...")
    df = pd.read_excel(config.RAW_DATA_PATH, sheet_name=0)
    print(f"Raw data loaded: {len(df)} rows, {len(df.columns)} columns")

    # --- Write manual inputs from config ---
    # These values come from other systems (ops, finance, complaints) and
    # cannot be derived from the survey data. They're set in config.py.
    print("Writing manual inputs from config...")

    pax_row = find_summary_row(ws_summary, "No. of Pax")
    ws_summary.cell(row=pax_row, column=target_col).value = config.PASSENGER_COUNT

    survey_row = find_summary_row(ws_summary, "No. of surveys achieved ")
    ws_summary.cell(row=survey_row, column=target_col).value = len(df)
    print(f"    - Survey count: {len(df)}")

    # --- Run all metric calculations ---
    for question in QUESTION_REGISTRY:
        print(f"Calculating: {question['name']}...")
        try:
            question["fn"](df, ws_summary, target_col)
        except Exception as e:
            print(f"  ERROR in {question['name']}: {e}")
            # We log and continue — one broken metric should not stop the rest

    # --- Save ---
    # Always saves to OUTPUT_PATH, never INPUT_PATH.
    # This means the original file is always preserved as a fallback.
    print(f"Saving to {config.OUTPUT_PATH}...")
    wb.save(config.OUTPUT_PATH)
    print("Done.")